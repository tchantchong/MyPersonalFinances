# Imports the Google Cloud client library
from google.cloud import language
from lists import *
from fin_lists import *
import openpyxl

# Instantiates a client
language_client = language.Client()

# Lists
words_list = words_list()
weights_list = weights_list()
fin_words_list = fin_words_list()
fin_wights_list = fin_weights_list()


def main():
    workbook = openpyxl.load_workbook('Crawler Results.xlsx')
    sheet = workbook.get_sheet_by_name('Textos Raw')
    symbol_list = []
    date_list = []
    title_list = []
    title_scores = []
    summary_list = []
    summary_scores = []
    analysis_list = []
    analysis_scores = []
    for i in range (2, sheet.max_row+1):
        symbol_list.append(sheet.cell(row=i, column=1).value)
        #date_list.append(sheet.cell(row=i, column=4).value)
        #title_list.append(sheet.cell(row=i, column=3).value)
        #title_scores.append(analyze_text(sheet.cell(row=i, column=3).value))
        summary_list.append(sheet.cell(row=i, column=6).value)
        summary_scores.append(analyze_text(sheet.cell(row=i, column=6).value))
        print ('Sumario: ', sheet.cell(row=i, column=6).value)
        print ('Score: ', summary_scores[i-2])
        #analysis_list.append(sheet.cell(row=i, column=7).value)
        #analysis_scores.append(analyze_text(sheet.cell(row=i, column=7).value))
    stock_list = list(set(symbol_list)) #symbol_list without duplicates
    #stock_title_scores = [0] * len(stock_list)
    stock_summary_scores = [0] * len(stock_list)
    #stock_analysis_scores = [0] * len(stock_list)
    
    for symbol in symbol_list:
        for stock in stock_list:
            if symbol == stock:
                #stock_title_scores[stock_list.index(stock)] += title_scores[symbol_list.index(symbol)]
                stock_summary_scores[stock_list.index(stock)] += summary_scores[symbol_list.index(symbol)]
                #stock_analysis_scores[stock_list.index(stock)] += analysis_scores[symbol_list.index(symbol)]

    print_final_scores(stock_list, stock_title_scores, stock_summary_scores, stock_analysis_scores)

    
def analyze_text(text):    # The text to analyze
    #text = 'The analysis provided in this article has not found any upside potential in Petroleo Brasileiro SA Petrobras\'s (NYSE:PBR) shares in the foreseeable future. Although the company has demonstrated promising financial results in the last quarter, there is a strong necessity to improve the management of working capital. Furthermore, Petroleo has a significant level of debt compared to peers. Another critical weakness is that the company does not change amid the global energy transition. We consider these factors as a headwind for revenue growth and long-term success.'
    sentences = split_text(text)
    sentences_rank = [0] * len(sentences)
    google_score = []
    google_magnitude = []
    text_score = 0
    print('TEXT: ', text)
    for sentence in sentences:
        document = language_client.document_from_text(sentence)
        annotations = document.annotate_text()
        print('Sentence: ', sentence)
        #print('Sentiment: {}, {}'.format(document.analyze_sentiment().sentiment.score, document.analyze_sentiment().sentiment.magnitude))
        google_score.append(document.analyze_sentiment().sentiment.score)
        google_magnitude.append(document.analyze_sentiment().sentiment.magnitude)
        
        #Refresh ranking with words matching to the library defined
        sentences_rank[sentences.index(sentence)] += sentence_matching_words(document, words_list, weights_list, fin_words_list, fin_weights_list)

        #Refresh ranking with correct dependencies
        

        #Refresh ranking with sentences tenses
        if (sentences_rank[sentences.index(sentence)] > 0):
            sentences_rank[sentences.index(sentence)]+=sentence_tense(document)
        elif (sentences_rank[sentences.index(sentence)] < 0):
            sentences_rank[sentences.index(sentence)]-=sentence_tense(document)
                
        for token in annotations.tokens:
            msg = '%11s %11s: %s' % (token.part_of_speech.tag, token.part_of_speech.tense, token.text_content)
            #print(msg)

        text_score += sentences_rank[sentences.index(sentence)]
        
    #Sanity check with Google Natural Language
    #print_google_check(sentences_rank, google_magnitude, google_score, False)


def split_text(text): #split a text in phrases with at least one verb
    document = language_client.document_from_text(text)
    annotations = document.annotate_text()
    sentences = []
    actual_sentence = ''
    verb_presence = False
    for token in annotations.tokens:
        if token.part_of_speech.tag == 'VERB':
            verb_presence = True
            if (actual_sentence == ''):
                actual_sentence += token.text_content
            else:
                actual_sentence = actual_sentence + ' ' + token.text_content
        elif is_punctuation(token.text_content):
            actual_sentence += token.text_content
            if verb_presence == True:
                sentences.append(actual_sentence)
                actual_sentence=''
                verb_presence = False
        else:
            if (actual_sentence == ''):
                actual_sentence += token.text_content
            else:
                actual_sentence = actual_sentence + ' ' + token.text_content             
    return sentences

def is_punctuation(word):
    if (word == '.' or word == '!' or word == '?' or word == ';' or word == ','):
        return True
    else:
        return False


def sentence_matching_words(sentence, words_list, weights_list, fin_words_list, fin_weights_list):
    sentence_score = 1
    key_words = []
    for token in sentence.annotate_text().tokens:
        for word in words_list:
            if token.text_content == word:
                sentence_score = sentence_score * weights_list[words_list.index(word)]
                if weights_list[words_list.index(word)] != 1:
                    key_words.append(word)
                    print(word + ': ' + str(weights_list[words_list.index(word)]))
    if sentence_score > 20:
        sentence_score = 20
    elif sentence_score < -20:
        sentence_score = -20

    #Rules for score correction:
    number_fin_words = 0
    for word in key_words:
        for fin_word in fin_words_list:
            if word == fin_word:
                number_fin_words += 1
    if (number_fin_words == len(key_words) or number_fin_words == 0):
        sentence_score = 0

    print ('Sentence score: ' + str(sentence_score))
    return sentence_score
    

#def sentence_correct_dependencies(sentence):


def sentence_tense(sentence):
    counter = 0
    for token in sentence.annotate_text().tokens:
        if token.part_of_speech.tense == 'PAST':
            counter -= 1
        elif token.part_of_speech.tense == 'PRESENT':
            counter += 0
        elif token.part_of_speech.tense == 'FUTURE':
            counter += 1
    if counter > 0:
        return (2) #additional score to future sentences
    elif counter == 0:
        return (2) #additional score to present sentences
    else:
        return (-2) #additional score to past sentences
    
def print_google_check(sentences_rank, google_scores, google_magnitudes, details):
    print ('=' * 84)
    print ('|' + ' SENTENCE ' + '|' + ' MATCHING WORDS SCORE ' + '|' + ' GOOGLE SCORE ' + '|' + ' GOOGLE MAGNITUDE ' + '|' + ' SANITY CHECK ' + '|' )
    match_sco_tot = 0
    google_sco_tot = 0
    google_magn_tot = 0
    for count in range (len(sentences_rank)):
        sen_num = count + 1
        match_sco = sentences_rank[count]
        google_sco = google_scores[count]
        google_magn = google_magnitudes[count]
        if ((match_sco >= 0 and google_sco >= 0) or (match_sco <= 0 and google_sco <= 0)):
            check = 'CHECK'
        else:
            check = 'NOT CHECK'
        if details == True:
            print ('|' + '----------' + '+' + '----------------------' + '+' + '--------------' + '+' + '------------------' + '+' + '--------------' + '|' )
            print ('|' + (' ' * (6-len(str(sen_num)))) + str(sen_num) + '    ' + '|' + (' ' * (12-len(str(match_sco)))) + str(match_sco) + '          ' + '|' + (' ' * (9-len(str(google_sco)))) + str(google_sco) + '     ' + '|' + (' ' * (12-len(str(google_magn)))) + str(google_magn) + '      ' + '|' + (' ' * (11-len(str(check)))) + str(check) + '   ' + '|' )
        match_sco_tot += match_sco
        google_sco_tot += google_sco
        google_magn_tot += google_magn
    sen_num = 'ALL'
    match_sco = match_sco_tot
    google_sco = google_sco_tot
    google_magn = google_magn_tot
    if ((match_sco >= 0 and google_sco >= 0) or (match_sco <= 0 and google_sco <= 0)):
        check = 'CHECK'
    else:
        check = 'NOT CHECK'
    print ('|' + '----------' + '+' + '----------------------' + '+' + '--------------' + '+' + '------------------' + '+' + '--------------' + '|' )
    print ('|' + (' ' * (6-len(str(sen_num)))) + str(sen_num) + '    ' + '|' + (' ' * (12-len(str(match_sco)))) + str(match_sco) + '          ' + '|' + (' ' * (9-len(str(google_sco)))) + str(google_sco) + '     ' + '|' + (' ' * (12-len(str(google_magn)))) + str(google_magn) + '      ' + '|' + (' ' * (11-len(str(check)))) + str(check) + '   ' + '|' )
    print ('=' * 84)

def print_final_scores(stocks, titles, summaries, analyses):
    print ('=' * 84)
    print ('|' + '   STOCK  ' + '|' + '     TITLE SCORE      ' + '|' + 'SUMMARY SCORE ' + '|' + '  ANALYSIS SCORE  ' + '|' )    
    for stock in stocks:
        tit = titles[stocks.index(stock)]
        summ = summaries[stocks.index(stock)]
        anal = analyses[stocks.index(stock)]
        print ('|' + '----------' + '+' + '----------------------' + '+' + '--------------' + '+' + '------------------' + '|' )
        print ('|' + (' ' * (6-len(str(stock)))) + str(stock) + '    ' + '|' + (' ' * (12-len(str(tit)))) + str(tit) + '          ' + '|' + (' ' * (9-len(str(summ)))) + str(summ) + '     ' + '|' + (' ' * (12-len(str(anal)))) + str(anal) + '      ' + '|')
    print ('=' * 84)
    
        
main()

